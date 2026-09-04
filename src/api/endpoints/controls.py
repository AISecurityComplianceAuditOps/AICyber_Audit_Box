import re
from fastapi import APIRouter, HTTPException, Query, File, UploadFile, Request, Form
from pydantic import BaseModel, Field, field_validator
from typing import List, Optional
from src.db.database import (
    get_all_custom_controls,
    add_custom_control,
    update_custom_control,
    delete_custom_control
)
from src.ai.keyword_generator import generate_keywords
from src.api.endpoints.auth import _require_auth
from src.core.text_validation import clean_safe_text as _clean_safe_text, clean_keywords as _clean_keywords

router = APIRouter(prefix="/controls", tags=["Manage Controls"])

# --- Input validation helpers ---
# These fields are stored and later rendered back into the UI (control table,
# framework accordion, report exports). The frontend HTML-escapes on render, but
# rejecting HTML/script-bearing input at the API boundary is the actual security
# gate -- it's the only place that's guaranteed to run regardless of which UI path
# (or a direct API call) is used to create/edit a control. Shared with other
# endpoint files via src/core/text_validation.py.
_CONTROL_ID_RE = re.compile(r'^[A-Za-z0-9][A-Za-z0-9 ._\-/()]*$')

# --- Framework support ---
# "VAPT" is the only framework with real backend behavior today: audit_graph.py
# carries this value through as AuditState["standard"], and audit_chains.py's
# is_vapt check swaps in CVSS-style prompting/severity when standard=="VAPT" --
# the same mechanism the 15 built-in VAPT-* controls already use. The others
# (DPDP/GDPR/SOC2/BCMS/XBOM) have no dedicated prompt template yet, so they fall
# back to the same generic evidence-based evaluation ISO 27001 controls get --
# that's an intentional, correct fallback, not a gap.
#
# For non-ISO frameworks we also rewrite `category` to a "<Framework> Framework
# Controls" string, matching exactly how the built-in VAPT entries are tagged
# ("VAPT Framework Controls") -- this is what makes the control show up under
# the matching "Target Framework" filter in app.js (cat.includes("VAPT")/
# "DPDP"/"SOC"/"BCMS"/"X-BOM"), with zero frontend filter-logic changes needed.
FRAMEWORK_CHOICES = ("ISO 27001", "VAPT", "DPDP", "GDPR", "SOC2", "BCMS", "XBOM", "PQC")
_FRAMEWORK_CATEGORY_LABELS = {
    "VAPT": "VAPT Framework Controls",
    "DPDP": "DPDP Framework Controls",
    "GDPR": "GDPR Framework Controls",
    "SOC2": "SOC 2 Framework Controls",
    "BCMS": "BCMS Framework Controls",
    "XBOM": "X-BOM / SBOM Framework Controls",
    "PQC": "PQC Framework Controls",
}

def _resolve_category_for_framework(framework: str, category: str) -> str:
    """ISO 27001 keeps the auditor-picked clause category as-is; every other
    framework gets a fixed, filter-matching category string (see above)."""
    return _FRAMEWORK_CATEGORY_LABELS.get(framework, category) if framework != "ISO 27001" else category

# --- Request / Response Schemas ---
class CreateControlRequest(BaseModel):
    control_id: str
    control_name: str
    category: str
    framework: str = "ISO 27001"
    keywords: List[str] = []
    description: str = ""
    is_global: bool = True
    created_by: str = "auditor"

    @field_validator("control_id")
    @classmethod
    def _validate_control_id(cls, v):
        v = _clean_safe_text(v, "Control ID", 40, required=True)
        if not _CONTROL_ID_RE.match(v):
            raise ValueError("Control ID may only contain letters, numbers, spaces, and . _ - / ( )")
        return v

    @field_validator("control_name")
    @classmethod
    def _validate_control_name(cls, v):
        return _clean_safe_text(v, "Control Name", 200, required=True)

    @field_validator("category")
    @classmethod
    def _validate_category(cls, v):
        return _clean_safe_text(v, "Category", 100)

    @field_validator("framework")
    @classmethod
    def _validate_framework(cls, v):
        v = (v or "ISO 27001").strip()
        if v not in FRAMEWORK_CHOICES:
            raise ValueError(f"Framework must be one of: {', '.join(FRAMEWORK_CHOICES)}")
        return v

    @field_validator("description")
    @classmethod
    def _validate_description(cls, v):
        return _clean_safe_text(v, "Description", 2000)

    @field_validator("keywords")
    @classmethod
    def _validate_keywords(cls, v):
        return _clean_keywords(v)

    @field_validator("created_by")
    @classmethod
    def _validate_created_by(cls, v):
        return _clean_safe_text(v, "Created By", 100) or "auditor"

class UpdateControlRequest(BaseModel):
    keywords: Optional[List[str]] = None
    description: Optional[str] = None
    is_active: Optional[bool] = None
    framework: Optional[str] = None
    category: Optional[str] = None

    @field_validator("description")
    @classmethod
    def _validate_description(cls, v):
        if v is None:
            return v
        return _clean_safe_text(v, "Description", 2000)

    @field_validator("keywords")
    @classmethod
    def _validate_keywords(cls, v):
        if v is None:
            return v
        return _clean_keywords(v)

    @field_validator("framework")
    @classmethod
    def _validate_framework(cls, v):
        if v is None:
            return v
        v = v.strip()
        if v not in FRAMEWORK_CHOICES:
            raise ValueError(f"Framework must be one of: {', '.join(FRAMEWORK_CHOICES)}")
        return v

    @field_validator("category")
    @classmethod
    def _validate_category(cls, v):
        if v is None:
            return v
        return _clean_safe_text(v, "Category", 100)

class AutogenKeywordsRequest(BaseModel):
    name: str
    description: str = ""

# --- Endpoints ---

@router.get("")
def api_get_controls(request: Request, active_only: bool = Query(True)):
    _require_auth(request)
    try:
        controls = get_all_custom_controls(active_only=active_only)
        return {"success": True, "controls": controls}
    except Exception:
        raise HTTPException(status_code=500, detail="Failed to load controls.")

@router.post("")
def api_create_control(request: Request, req: CreateControlRequest):
    user = _require_auth(request)
    # These controls default to is_global=True (shared across every auditor), so
    # any authenticated role including auditee could otherwise create/modify
    # controls used by everyone -- matching the role gate DELETE already enforces
    # below.
    if user.get("role") not in ("admin", "auditor"):
        raise HTTPException(status_code=403, detail="Access denied.")
    if not req.control_id.strip() or not req.control_name.strip():
        raise HTTPException(status_code=400, detail="Control ID and Control Name are required.")
    try:
        resolved_category = _resolve_category_for_framework(req.framework, req.category.strip())
        new_id = add_custom_control(
            control_id=req.control_id.strip(),
            control_name=req.control_name.strip(),
            category=resolved_category,
            framework=req.framework,
            keywords=req.keywords,
            description=req.description.strip(),
            auto_generated=False,
            created_by=req.created_by.strip(),
            is_global=req.is_global
        )
        return {"success": True, "message": "Control saved successfully", "id": new_id}
    except Exception:
        raise HTTPException(status_code=500, detail="Failed to save control.")

@router.post("/autogen-keywords")
def api_autogen_keywords(request: Request, req: AutogenKeywordsRequest):
    _require_auth(request)
    if not req.name.strip():
        raise HTTPException(status_code=400, detail="Control name is required for keyword generation.")
    try:
        keywords = generate_keywords(req.name.strip(), req.description.strip())
        return {"success": True, "keywords": keywords}
    except Exception:
        raise HTTPException(status_code=500, detail="Keyword generation failed.")

@router.put("/{db_id}")
def api_update_control(request: Request, db_id: int, req: UpdateControlRequest):
    user = _require_auth(request)
    if user.get("role") not in ("admin", "auditor"):
        raise HTTPException(status_code=403, detail="Access denied.")
    try:
        # Switching to a non-ISO framework always forces the matching fixed
        # category label (the ISO clause dropdown value doesn't apply there);
        # leaving framework unset or ISO 27001 passes category through as given.
        resolved_category = req.category
        if req.framework is not None and req.framework != "ISO 27001":
            resolved_category = _FRAMEWORK_CATEGORY_LABELS.get(req.framework, req.category)
        success = update_custom_control(
            control_db_id=db_id,
            keywords=req.keywords,
            description=req.description,
            is_active=req.is_active,
            framework=req.framework,
            category=resolved_category
        )
        if not success:
            raise HTTPException(status_code=404, detail="Control not found in database.")
        return {"success": True, "message": "Control updated successfully"}
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=500, detail="Failed to update control.")

@router.delete("/{db_id}")
def api_delete_control(request: Request, db_id: int, soft: bool = Query(True)):
    user = _require_auth(request)
    if user.get("role") not in ("admin", "auditor"):
        raise HTTPException(status_code=403, detail="Access denied.")
    try:
        success = delete_custom_control(control_db_id=db_id, soft=soft)
        if not success:
            raise HTTPException(status_code=404, detail="Control not found in database.")
        action_type = "deactivated" if soft else "deleted"
        return {"success": True, "message": f"Control successfully {action_type}"}
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=500, detail="Failed to delete control.")

@router.get("/framework")
def api_get_framework_controls(request: Request):
    _require_auth(request)
    """Returns combined list of standard ISO/VAPT framework controls and custom controls."""
    try:
        from src.core.controls_data import USE_CASES
        from src.core.bg_worker import _load_custom_use_cases
        customs = _load_custom_use_cases(force=True)
        combined = []
        for uc in USE_CASES:
            combined.append({"sl": uc["sl"], "use_case": uc["use_case"], "label": uc["label"], "category": uc["category"]})
        for c in customs:
            combined.append({"sl": c["sl"], "use_case": c["use_case"], "label": c["label"], "category": c["category"]})
        return {"success": True, "controls": combined}
    except Exception:
        raise HTTPException(status_code=500, detail="Failed to load framework controls.")


def _build_scope_payload(items):
    """Resolve parsed checklist rows to ISO controls and build the scope payload.

    Shared by the Excel upload endpoint and the in-app checklist builder so both
    entry points resolve controls identically -- the resolution here (exact id,
    label prefix, keyword map, name match, word overlap) runs AFTER the parser and
    is what actually decides which controls a checklist maps to. Duplicating it for
    the builder would have let the two paths drift apart silently.
    """
    # Imported here rather than at module scope: this used to live inside the
    # upload endpoint, which imported it locally, and controls_data pulls in the
    # full 217-control table that not every route in this file needs.
    from src.core.controls_data import USE_CASES

    custom_evidence = {}
    custom_documents = {}
    # BUG FIX: Use LIST not SET — two rows with same ctrl_id (e.g. two 8.17 NTP checks)
    # must each appear so progress counter and control loop both get 8 items, not 6.
    matched_sls_list = []
    import re as _re
    from src.core.excel_scoping_parser import (
        _resolve_control_by_direct_map as _kw_resolve,
        _resolve_control_by_name as _name_resolve,
    )

    def _norm(t: str) -> str:
        """Lowercase + strip punctuation for fuzzy comparisons."""
        return _re.sub(r'[^a-z0-9\s]', ' ', str(t or '').lower()).strip()

    for item in items:
        ctrl_id = item.get("control_id")
        ctrl_label = item.get("control_label")
        question = item.get("question") or item.get("control_label") or ctrl_id
        expected_ev = item.get("expected_evidence") or question or ""
        files = item.get("files") or item.get("raw_file_refs") or []
        files_str = ", ".join(files) if isinstance(files, list) else str(files)

        matched_uc = None

        # --- Pass 1: exact numeric ID match or exact use_case string match ---
        for uc in USE_CASES:
            uc_id = uc["use_case"].split(" ")[0]
            if uc_id == ctrl_id or uc["use_case"] == ctrl_label:
                matched_uc = uc
                break

        # --- Pass 2: strip leading numeric ID from ctrl_label and compare ---
        # Handles the case where the parser resolved via keyword/embedding and
        # ctrl_id == "UNKNOWN" but ctrl_label == "5.15 Access Control" (a valid
        # ISO label whose numeric prefix still uniquely identifies the control).
        if not matched_uc and ctrl_label and ctrl_label != "UNKNOWN":
            label_id_match = _re.match(r'^(VAPT\s*-?\s*\d+|\d{1,2}\.\d{1,2}(?:\.\d{1,2})?)', str(ctrl_label).strip(), _re.IGNORECASE)
            if label_id_match:
                label_prefix = label_id_match.group(1).strip().upper()
                for uc in USE_CASES:
                    uc_id = uc["use_case"].split(" ")[0].upper()
                    if uc_id == label_prefix:
                        matched_uc = uc
                        break

        # --- Pass 2.5: re-run the parser's rich keyword map on the question ---
        # The parser already ran _resolve_control_by_direct_map on resolution_text
        # but returned UNKNOWN when that text was empty or the question column was
        # detected differently from what the backend stores in item["question"].
        # Re-running here covers question-based rows like "Is MFA enabled?" or
        # "Are backups taken daily?" that carry domain keywords (mfa -> 8.5,
        # backup -> 8.13, ntp -> 8.17) that the word-overlap pass below would miss
        # because it only compares against USE_CASES label words, not audit terms.
        if not matched_uc and question:
            matched_uc = _kw_resolve(question, USE_CASES)

        # Also try _resolve_control_by_name on the question text (handles cases
        # like "Clock Synchronization" or "Backup and Recovery" as plain text).
        if not matched_uc and question:
            matched_uc = _name_resolve(question, USE_CASES)

        # --- Pass 3: word-overlap fallback on ctrl_label text ---
        # Final resort when neither ID, label-prefix, nor keyword/name pass matched
        # (e.g. a highly domain-specific or abbreviated question).
        if not matched_uc:
            search_text = _norm(" ".join(filter(None, [ctrl_label, question])))
            search_words = [w for w in search_text.split() if len(w) > 2]
            if search_words:
                best_uc, best_score = None, 0
                for uc in USE_CASES:
                    uc_text = _norm(uc.get("label", "") + " " + uc.get("use_case", ""))
                    uc_words = set(uc_text.split())
                    overlap = len(set(search_words) & uc_words)
                    if overlap > best_score:
                        best_score, best_uc = overlap, uc
                min_req = 1 if len(search_words) <= 2 else 2
                if best_score >= min_req:
                    matched_uc = best_uc

        if matched_uc:
            uc_key = matched_uc["use_case"]
            uc_id = uc_key.split(" ")[0]

            # A row is allowed to start with NO control id (Customize's whole
            # point: "leave it blank, the question is matched automatically"),
            # and this loop resolves it correctly for MATCHING purposes above --
            # but until now, only for matching. The resolved id/name was never
            # written back onto the row itself, so a row that started blank
            # STAYED blank in `items` (stored verbatim into excel_items below),
            # even though the right control had genuinely been found. The
            # finding built from that row then had no control_id/control_name
            # at all, and the UI's header rendering, with nothing before the
            # " — " it inserts, showed a bare "— <question>" -- Control ID and
            # Control Name both blank, only the question visible.
            if not ctrl_id or ctrl_id == "UNKNOWN":
                item["control_id"] = uc_id
                ctrl_id = uc_id
            if not ctrl_label or ctrl_label == "UNKNOWN":
                item["control_label"] = uc_key
                ctrl_label = uc_key

            # BUG FIX: append to list so two 8.17 rows both appear
            matched_sls_list.append(int(matched_uc["sl"]))

            # BUG FIX: merge files for same ctrl_id rather than overwrite.
            # This means both NTP evidence files end up in the mapping for 8.17.
            for target_k in (uc_key, uc_id, ctrl_label):
                if target_k:
                    if target_k not in custom_documents or not custom_documents[target_k]:
                        custom_documents[target_k] = files_str
                    elif files_str and files_str not in custom_documents[target_k]:
                        # Append additional file for same control, comma-separated
                        custom_documents[target_k] = custom_documents[target_k] + ", " + files_str
                    if expected_ev:
                        custom_evidence[target_k] = expected_ev

    custom_evidence["excel_items"] = items

    unmatched_count = len(items) - len(matched_sls_list)
    warning = None
    if len(items) > 0 and len(matched_sls_list) == 0:
        warning = (
            "No ISO controls could be matched from the Excel sheet. "
            "Please ensure the sheet has a column containing ISO control IDs (e.g. '5.15') "
            "or recognisable control names."
        )
    elif unmatched_count > 0:
        warning = f"{unmatched_count} row(s) could not be matched to a known ISO control and were skipped."

    return {
        "success": True,
        "matched_sls": matched_sls_list,        # list with duplicates — 8 items, not 6
        "custom_evidence": custom_evidence,
        "custom_documents": custom_documents,
        "total_rows": len(items),
        "warning": warning,
        "message": f"Successfully loaded {len(items)} Excel audit checklist items ({len(set(matched_sls_list))} unique ISO controls)."
    }

class ChecklistRow(BaseModel):
    """One row of the in-app checklist builder.

    `files` is the evidence side. `policy_files` is the policy side, and exists
    because Excel Scoping assesses the two separately -- the builder could only
    ever produce evidence, so a scope built in-app was structurally incapable of
    passing an Excel-mode row no matter what the auditor typed.
    """
    question: str = ""
    files: List[str] = []
    policy_files: List[str] = []
    control_id: str = ""

    @field_validator("question")
    @classmethod
    def _validate_question(cls, v):
        return _clean_safe_text(v, "Audit check question", 500)

    @field_validator("control_id")
    @classmethod
    def _validate_ctrl_id(cls, v):
        return _clean_safe_text(v, "Control ID", 40)

    @field_validator("files")
    @classmethod
    def _validate_files(cls, v):
        # Filenames are chosen from the session's own uploaded evidence in the UI,
        # but this endpoint is reachable directly, so they get the same free-text
        # treatment as every other persisted string in this file.
        return [_clean_safe_text(f, "File name", 260) for f in (v or []) if str(f or "").strip()]


class BuildChecklistRequest(BaseModel):
    rows: List[ChecklistRow] = []
    scoping_mode: str = ""


@router.post("/build-scope-checklist")
def api_build_scope_checklist(request: Request, req: BuildChecklistRequest):
    """Build a scope from checklist rows typed into the app, with no Excel file.

    Produces exactly the same payload as /parse-scope-excel by feeding the rows
    through the same _build_scope_payload() resolution, so an in-app checklist and
    an uploaded sheet are indistinguishable downstream -- same control matching,
    same custom_evidence shape, same Customize handling.

    The UI picks filenames from the evidence already uploaded to the session, which
    removes the whole class of "typed filename doesn't match the uploaded file"
    mismatches that the Excel path has to fuzzy-match its way around.
    """
    _require_auth(request)
    customize = str(req.scoping_mode or "").strip().upper().startswith("CUSTOM")

    # A Customize row IS its question, so an empty one is nothing. An Excel row
    # is a control plus the documents to assess, and requiring a question there
    # discarded every row the auditor filled in correctly -- the mode does not
    # ask for one.
    if customize:
        rows = [r for r in (req.rows or []) if (r.question or "").strip()]
        if not rows:
            raise HTTPException(status_code=400, detail="Add at least one audit check question.")
    else:
        rows = [r for r in (req.rows or [])
                if (r.control_id or "").strip() or (r.files or []) or (r.policy_files or [])]
        if not rows:
            raise HTTPException(
                status_code=400,
                detail="Add at least one row with a control ID and the policy or evidence document to assess.")

    # Resolve a typed control ID exactly as an uploaded sheet's ID column is
    # resolved. _build_scope_payload's own passes compare the id verbatim against
    # "6.5" and match a label only when it STARTS with a digit, so the ISO Annex A
    # form -- "A.6.5", the form printed on the standard, used in the customer's own
    # checklist, and shown in this field's placeholder -- matched nothing at all:
    # the row was accepted, reported as built, and put zero controls in scope.
    # Reusing the parser's resolver means any id shape the spreadsheet accepts,
    # the builder accepts too.
    try:
        from src.core.controls_data import USE_CASES as _ALL_UCS
        from src.core.excel_scoping_parser import _resolve_control_by_id
    except Exception:
        _ALL_UCS, _resolve_control_by_id = [], None

    def _canonical_control(raw_id):
        """(control_id, control_label, expected_evidence) for a typed id."""
        raw = (raw_id or "").strip()
        if not raw or not _resolve_control_by_id:
            return raw, raw, ""
        try:
            uc = _resolve_control_by_id(raw, _ALL_UCS)
        except Exception:
            uc = None
        if not uc:
            return raw, raw, ""
        label = str(uc.get("use_case", ""))
        return label.split(" ", 1)[0], label, str(uc.get("expected", ""))

    items = []
    for idx, r in enumerate(rows):
        q = (r.question or "").strip()
        _cid, _clabel, _cexpected = _canonical_control(r.control_id)
        evidence_files = list(r.files or [])
        # Customize has no policy dimension by design; anything sent for one in
        # that mode is dropped rather than quietly reintroducing the policy rule
        # the mode exists to waive.
        policy_files = [] if customize else list(r.policy_files or [])
        files = policy_files + [f for f in evidence_files if f not in policy_files]
        items.append({
            "row_index": idx + 1,
            "question": q,
            "requirement_question": q,
            "requirement_question_source": "builder",
            "requirement_question_status": "RESOLVED" if q else "UNRESOLVED",
            "files": files,
            "policy_files": policy_files,
            "evidence_files": evidence_files,
            "raw_file_refs": files,
            # Resolved above where the auditor named one; otherwise left for
            # _build_scope_payload() to work out, exactly as an Excel row with no
            # control column would be.
            "control_id": _cid,
            "control_name": "",
            "control_label": _clabel,
            "expected_evidence": _cexpected,
            "prompt_hint": "",
            "severity": "MEDIUM",
            "customize_mode": customize,
            "policy_required": not customize,
        })

    payload = _build_scope_payload(items)
    _shape = "question" if customize else "policy/evidence"
    payload["message"] = (
        f"Built {len(items)} {_shape} checklist item(s) from the in-app builder "
        f"({len(set(payload['matched_sls']))} unique control(s) matched)."
    )
    return payload


@router.get("/scope-template")
def api_scope_template(request: Request, mode: str = Query("EXCEL")):
    """Download a blank checklist template in the shape THIS scope mode reads.

    One template used to serve both modes, and it was the question-based one:
    "Audit check | File name | File type | Control ID". That is the Customize
    shape. Excel Scoping does not judge a question -- it assesses a POLICY
    document and an EVIDENCE document against the control, and a row is only
    compliant when BOTH hold up. So an auditor working in Excel mode downloaded
    a sheet with nowhere to name the policy, filled it in, uploaded it, and got
    question-based scoping out of the one mode whose entire purpose is the
    policy/evidence split -- while the sheet's own "How to use" page told them
    the same sheet served both modes.

    The parser reads either shape; this only stops the app handing out the
    wrong one.
    """
    _require_auth(request)
    try:
        import io as _io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from fastapi.responses import StreamingResponse

        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Checklist"

        _customize = str(mode or "").strip().upper().startswith("CUSTOM")

        if _customize:
            headers = ["Audit check", "File name", "File type", "Control ID (optional)"]
            samples = [
                ["Whether NTP is enabled and synchronized?", "121_NTP_Server_Clock_Sync.jpg", "JPG", ""],
                ["How is authentication implemented?", "Authentication_remark.txt", "TXT", "8.5"],
                ["Whether log archival is done?", "117_Log_Archived_Prod.jpg", "JPG", ""],
            ]
            widths = (58, 38, 12, 22)
            notes_lines = [
                ["CUSTOMIZE (QUESTION-BASED) CHECKLIST TEMPLATE"],
                [""],
                ["1. 'Audit check'  - your question, one per row. Required."],
                ["2. 'File name'    - the evidence file that answers it. Must match a file"],
                ["                    you upload as evidence. Separate several with a comma."],
                ["3. 'File type'    - optional, for your own reference only."],
                ["4. 'Control ID'   - optional. Leave blank and the question is matched to a"],
                ["                    control automatically; fill it in to pin the row."],
                [""],
                ["This is the CUSTOMIZE sheet. Each row is judged on whether the cited"],
                ["evidence answers the question. No policy document is required, and a row"],
                ["is never failed for missing one."],
                [""],
                ["For Excel Scoping, download the template again with Excel Scoping"],
                ["selected -- that mode reads a different sheet (policy + evidence)."],
            ]
        else:
            headers = ["Control ID (ISO)", "Policy Document Name", "Evidence Document Name"]
            samples = [
                ["A.6.5", "Access_Control_-_Organizational_Controls.docx",
                 "Hiring_and_Termination_Process_with_evidences.pptx"],
                ["A.5.15", "Access_Control_Policy.docx", "Quarterly_Access_Review_Q1_2026.xlsx"],
                ["8.15", "Logging_and_Monitoring_Policy.docx", "SIEM_Log_Retention_Evidence.png"],
            ]
            widths = (22, 52, 52)
            notes_lines = [
                ["EXCEL SCOPING CHECKLIST TEMPLATE"],
                [""],
                ["1. 'Control ID (ISO)'        - the control this row assesses, e.g. A.6.5"],
                ["                               or 6.5. Required."],
                ["2. 'Policy Document Name'    - the policy that states the requirement."],
                ["                               Must match a file you upload as evidence."],
                ["3. 'Evidence Document Name'  - the operational proof that it is actually"],
                ["                               done. Must match an uploaded file too."],
                [""],
                ["Separate several files in one cell with a comma."],
                [""],
                ["How a row is judged"],
                ["  Excel Scoping assesses the two documents SEPARATELY against the"],
                ["  control, and reports both:"],
                ["    POLICY   - is the requirement documented, current and approved?"],
                ["    EVIDENCE - does the operational proof show it being done?"],
                ["  The row is COMPLIANT only when BOTH hold. A policy with no evidence,"],
                ["  or evidence with no policy behind it, is a gap and is reported as one."],
                [""],
                ["Retrieval is locked to the files named on the row, so the model never"],
                ["sees unrelated evidence when judging this control."],
                [""],
                ["Leave the policy cell blank only if the control genuinely has no policy"],
                ["dimension -- otherwise the row is reported as a policy gap. If you want"],
                ["question-based scoping with no policy at all, use Customize mode and"],
                ["download that template instead."],
            ]

        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2563EB")
            cell.alignment = Alignment(vertical="center")

        # Examples, so the expected shape is obvious without reading any docs.
        for row in samples:
            ws.append(row)

        for col, width in zip("ABCDE", widths):
            ws.column_dimensions[col].width = width

        notes = wb.create_sheet("How to use")
        for line in notes_lines:
            notes.append(line)
        notes.column_dimensions["A"].width = 82
        notes["A1"].font = Font(bold=True, size=13)

        _tpl_name = ("customize_question_checklist_template.xlsx" if _customize
                     else "excel_scoping_policy_evidence_template.xlsx")
        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=\"" + _tpl_name + "\""},
        )
    except Exception as e:
        print(f"[SCOPE TEMPLATE ERROR] {e}", flush=True)
        raise HTTPException(status_code=500, detail="Failed to generate the checklist template.")


@router.post("/parse-scope-excel")
async def api_parse_scope_excel(request: Request, file: UploadFile = File(...),
                                framework: str = Form(None),
                                scoping_mode: str = Form(None)):
    """Parses an uploaded auditor scope Excel mapping (.xlsx/.xls) and returns mapped control SLs and custom evidence.

    `framework` confines control resolution to the standard being audited. A row
    that carries no control ID is otherwise matched by name or question against
    all 217 controls from all eight frameworks -- an ISO row reading
    "Cryptographic Controls" resolved to SOC 2 CC3.4 rather than ISO 8.24, and
    the auditor's evidence was then judged against the wrong requirement.
    Optional: omitted means the previous all-framework behaviour.
    """
    _require_auth(request)
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported.")

        
    try:
        import os
        import tempfile
        from src.core.excel_scoping_parser import parse_excel_scoping_checklist
        from src.core.controls_data import USE_CASES
        from src.core.input_guardrail import scan_document

        contents = await file.read()

        # Same 4-layer structural scan (magic bytes, VBA macros, zip-bomb
        # ratio/size, embedded dangerous extensions) applied to evidence
        # uploads -- this endpoint was writing straight to a tempfile and
        # parsing with no scan at all. .xlsx is itself a ZIP container, so
        # this is exactly the file type the zip-bomb/macro layers exist for.
        is_clean, reason = scan_document(file.filename, contents, "")
        if not is_clean:
            raise HTTPException(status_code=400, detail=f"File rejected by security scan: {reason}")

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(contents)
            tmp_path = tmp.name

        try:
            # Customize scope = question-based audit. Control resolution still runs
            # (the matched control is shown and evaluated); the flag only removes the
            # policy requirement from the verdict downstream.
            _customize = str(scoping_mode or "").strip().upper().startswith("CUSTOM")
            items = parse_excel_scoping_checklist(
                tmp_path, framework=framework, customize_mode=_customize
            )
        finally:
            try:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except Exception:
                pass


        return _build_scope_payload(items)
    except HTTPException:
        raise
    except Exception as e:
        print(f"[EXCEL SCOPE PARSE ERROR] {e}", flush=True)
        raise HTTPException(status_code=500, detail="Failed to parse Excel scope file. Please check the file format and try again.")


